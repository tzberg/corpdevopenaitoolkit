#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CorpDev AI Platform - Enhanced Backend with Module-Specific Prompts
"""

import os
import hashlib
import mimetypes
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from pathlib import Path
import json
import re

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import uvicorn

# Optional dependencies with graceful fallbacks
try:
    from openai import OpenAI

    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("OpenAI not available. Install with: pip install openai")

try:
    from PyPDF2 import PdfReader

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Configuration
MAX_UPLOAD_BYTES = 100 * 1024 * 1024  # 100MB
ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xlsx", ".txt", ".json"}
UPLOAD_DIR = Path("./uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global application state
app_state = {
    "uploads": {},
    "modules": {
        "target_sourcing": {
            "label": "Target Sourcing",
            "description": "Identify 30 potential acquisition targets",
            "icon": "fas fa-search"
        },
        "due_diligence": {
            "label": "Due Diligence",
            "description": "Generate DD framework and checklist",
            "icon": "fas fa-clipboard-check"
        },
        "valuation": {
            "label": "Valuation Analysis",
            "description": "DCF with base/upside/downside cases",
            "icon": "fas fa-calculator"
        },
        "market_analysis": {
            "label": "Market Analysis",
            "description": "Competition and market dynamics",
            "icon": "fas fa-chart-area"
        },
        "integration": {
            "label": "Integration Planning",
            "description": "Post-acquisition integration process",
            "icon": "fas fa-puzzle-piece"
        },
        "synergies": {
            "label": "Synergy Analysis",
            "description": "Revenue and cost synergy identification",
            "icon": "fas fa-handshake"
        }
    }
}


# Utility functions
def generate_hash(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()[:12]


def is_allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def extract_pdf_content(file_path: Path) -> str:
    if not PDF_AVAILABLE:
        return "PDF extraction unavailable"
    try:
        reader = PdfReader(str(file_path))
        text_parts = []
        for i, page in enumerate(reader.pages[:10], 1):
            try:
                text = page.extract_text() or ""
                if text.strip():
                    text_parts.append(f"--- Page {i} ---\n{text}")
            except Exception:
                continue
        return "\n".join(text_parts)[:10000]
    except Exception as e:
        logger.warning(f"PDF extraction error: {e}")
        return f"PDF extraction error: {e}"


def extract_csv_content(file_path: Path) -> str:
    try:
        if PANDAS_AVAILABLE:
            df = pd.read_csv(file_path)
            preview = df.head(10).to_string()
            return f"CSV Overview: {len(df)} rows × {len(df.columns)} columns\nColumns: {', '.join(df.columns)}\n\nPreview:\n{preview}"
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()[:20]
            return f"CSV Preview (first 20 lines):\n{''.join(lines)}"
    except Exception as e:
        return f"CSV extraction error: {e}"


def extract_xlsx_content(file_path: Path) -> str:
    if not XLSX_AVAILABLE:
        return "Excel extraction unavailable"
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        content = []
        for sheet_name in workbook.sheetnames[:3]:
            sheet = workbook[sheet_name]
            content.append(f"=== Sheet: {sheet_name} ===")
            for row in sheet.iter_rows(max_row=10, values_only=True):
                content.append(" | ".join(
                    str(cell) if cell is not None else "" for cell in row))
        return "\n".join(content)[:10000]
    except Exception as e:
        return f"Excel extraction error: {e}"


def extract_text_content(file_path: Path) -> str:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()[:10000]
        return content
    except Exception as e:
        return f"Text extraction error: {e}"


def extract_file_content(file_path: Path) -> str:
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf_content(file_path)
    elif ext == ".csv":
        return extract_csv_content(file_path)
    elif ext == ".xlsx":
        return extract_xlsx_content(file_path)
    elif ext in {".txt", ".json"}:
        return extract_text_content(file_path)
    else:
        return "Unsupported file type"


def call_openai_api(prompt: str, api_key: str, max_tokens: int = 3000) -> str:
    """Call OpenAI API with optimized settings for concise responses"""
    if not OPENAI_AVAILABLE:
        return "OpenAI not available. Please install the openai package."

    try:
        client = OpenAI(api_key=api_key)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "You are a senior M&A analyst with 15+ years of experience. Be direct, concise, and quantitative. No fluff or unnecessary explanations. Focus on actionable insights and specific numbers."
                },
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=max_tokens
        )
        return response.choices[0].message.content
    except Exception as e:
        logger.error(f"OpenAI API error: {e}")
        return f"AI analysis error: {str(e)}"


def build_analysis_prompt(module: str, form_data: dict,
                          documents: List[str]) -> str:
    """Build optimized prompts for each module with smart defaults"""

    def get_or_default(key, default="Not specified - use your best judgment"):
        val = form_data.get(key, '').strip()
        return val if val else default

    prompts = {
        "target_sourcing": f"""You are a senior M&A analyst. Provide EXACTLY 30 potential acquisition targets for {form_data.get('acquirer', 'the acquirer')}.

Context:
- Acquirer: {form_data.get('acquirer', 'Not specified')}
- Target Industry: {get_or_default('targetIndustry', 'Infer appropriate industries based on acquirer profile')}
- Geography: {get_or_default('geography', 'Search globally but prioritize markets where acquirer operates')}
- Revenue Range: {get_or_default('revenueRange', 'Suggest appropriate range based on acquirer size and typical deal profile')}
- Strategic Notes: {get_or_default('notes', 'None provided')}

Format as a numbered list (1-30) with:
Company Name | Industry | Location | Est. Revenue | Key Rationale (one line)

Example:
1. Acme Corp | SaaS | San Francisco | $25M ARR | Strong product-market fit in SMB segment

Be concise. No introduction, no conclusion. Just the 30 companies.""",

        "due_diligence": f"""Create a concise due diligence framework for acquiring {form_data.get('target', 'the target')}.

Context:
- Acquirer: {form_data.get('acquirer', 'Not specified')}
- Target: {form_data.get('target', 'Not specified')}
- Notes: {get_or_default('notes', 'None provided')}

Provide:

1. FINANCIAL DD
   - Revenue quality & growth trajectory
   - Unit economics & CAC/LTV
   - Working capital requirements
   - Key metrics to verify

2. COMMERCIAL DD
   - Customer concentration (top 10)
   - Churn analysis
   - Pipeline quality
   - Competitive positioning

3. OPERATIONAL DD
   - Tech stack & scalability
   - Key person dependencies
   - Org structure gaps

4. LEGAL DD
   - Material contracts
   - IP ownership
   - Litigation exposure

5. TOP 5 RISKS
   [List specific red flags to investigate]

6. TIMELINE
   Week 1-2: [activities]
   Week 3-4: [activities]
   Week 5-6: [activities]

Be direct. No fluff.""",

        "valuation": f"""Perform a quick valuation analysis for {form_data.get('target', 'the target')}.

Context:
- Acquirer: {form_data.get('acquirer', 'Not specified')}
- Target: {form_data.get('target', 'Not specified')}
- WACC: {form_data.get('wacc') + '%' if form_data.get('wacc', '').strip() else 'Use industry-standard WACC (~9% for typical tech/growth company, adjust based on target profile)'}
- Terminal Growth: {form_data.get('terminalGrowth') + '%' if form_data.get('terminalGrowth', '').strip() else 'Use standard terminal growth rate (~2.5% for mature markets)'}
- Tax Rate: {form_data.get('taxRate') + '%' if form_data.get('taxRate', '').strip() else 'Use standard corporate tax rate (~21% US federal)'}
- Notes: {get_or_default('notes', 'None provided')}

Provide:

1. KEY ASSUMPTIONS
   Revenue CAGR: [%]
   EBITDA Margin: [%]
   Capex as % Revenue: [%]
   NWC as % Revenue: [%]

2. THREE CASES
   BASE: [Revenue growth %, EBITDA margin %, other key assumptions]
   UPSIDE: [Revenue growth %, EBITDA margin %, other key assumptions]
   DOWNSIDE: [Revenue growth %, EBITDA margin %, other key assumptions]

3. COMPARABLE COMPANIES (show 5-7 with multiples)
   Company | EV/Revenue | EV/EBITDA | Growth Rate
   [List format]

4. PRECEDENT TRANSACTIONS (show 3-5 recent deals)
   Target | Acquirer | EV/Revenue | Date
   [List format]

5. VALUATION RANGE
   Method | Low | Base | High
   DCF | $XXM | $XXM | $XXM
   Trading Comps | $XXM | $XXM | $XXM
   Transaction Comps | $XXM | $XXM | $XXM

   Implied Equity Value: $XXM - $XXM
   Per Share (if applicable): $XX - $XX

Skip methodology explanations. Just numbers and ranges.""",

        "market_analysis": f"""Analyze the market for {form_data.get('targetMarket', form_data.get('target', 'the target market'))}.

Context:
- Acquirer: {form_data.get('acquirer', 'Not specified')}
- Target Market: {form_data.get('targetMarket', 'Not specified')}
- Sample Company: {get_or_default('sampleCompany', 'Select relevant public companies as benchmarks')}
- Notes: {get_or_default('notes', 'None provided')}

Provide:

1. MARKET SIZE
   TAM: $XXB
   SAM: $XXB
   SOM: $XXB
   CAGR (2024-2028): XX%

2. KEY COMPETITORS
   Company | Market Share | Revenue | Key Strength
   [List top 5-7 players in table format]

3. TAILWINDS (quantify where possible)
   • [Trend 1 with growth rate or $ impact]
   • [Trend 2 with growth rate or $ impact]
   • [Trend 3 with growth rate or $ impact]
   • [Trend 4 with growth rate or $ impact]

4. HEADWINDS (quantify where possible)
   • [Challenge 1 with potential $ impact]
   • [Challenge 2 with potential $ impact]
   • [Challenge 3 with potential $ impact]
   • [Challenge 4 with potential $ impact]

5. STRATEGIC IMPLICATIONS
   [2-3 bullet points on what this means for the deal]

Be quantitative. Keep it tight.""",

        "integration": f"""Create a post-acquisition integration plan for {form_data.get('target', 'the target')}.

Context:
- Acquirer: {form_data.get('acquirer', 'Not specified')}
- Target: {form_data.get('target', 'Not specified')}
- Notes: {get_or_default('notes', 'None provided')}

Provide:

1. DAY 1 CRITICAL PATH
   • [Must-have item 1 + Owner]
   • [Must-have item 2 + Owner]
   • [Must-have item 3 + Owner]
   • [Must-have item 4 + Owner]
   • [Must-have item 5 + Owner]

2. FIRST 100 DAYS
   Week 1-2: [Key milestones]
   Week 3-4: [Key milestones]
   Week 5-8: [Key milestones]
   Week 9-14: [Key milestones]

3. INTEGRATION TEAM STRUCTURE
   • Integration Lead: [Role/responsibilities]
   • Finance Workstream: [Lead + focus areas]
   • Operations Workstream: [Lead + focus areas]
   • Technology Workstream: [Lead + focus areas]
   • HR/Culture Workstream: [Lead + focus areas]

4. QUICK WINS (capture value fast)
   • [Win 1 with $ impact and timeline]
   • [Win 2 with $ impact and timeline]
   • [Win 3 with $ impact and timeline]

5. TOP 3 RISKS + MITIGATIONS
   Risk 1: [description] → Mitigation: [action]
   Risk 2: [description] → Mitigation: [action]
   Risk 3: [description] → Mitigation: [action]

6. SUCCESS METRICS
   • [KPI 1 with target]
   • [KPI 2 with target]
   • [KPI 3 with target]
   • [KPI 4 with target]

Process-focused. Clear owners and timelines.""",

        "synergies": f"""Identify and quantify synergies from acquiring {form_data.get('target', 'the target')}.

Context:
- Acquirer: {form_data.get('acquirer', 'Not specified')}
- Target: {form_data.get('target', 'Not specified')}
- Notes: {get_or_default('notes', 'None provided')}

Provide:

1. REVENUE SYNERGIES
   • Cross-sell existing products to target customers: $XXM
   • Upsell target products to acquirer base: $XXM
   • Geographic expansion: $XXM
   • New product bundles: $XXM
   Total Revenue Synergies: $XXM

2. COST SYNERGIES
   • Headcount reduction/consolidation: $XXM
   • Vendor/software consolidation: $XXM
   • Facilities/real estate: $XXM
   • G&A efficiencies: $XXM
   • Sales & marketing optimization: $XXM
   Total Cost Synergies: $XXM

3. SYNERGY QUANTIFICATION BY YEAR
   Year 1: $XXM (X% of total)
   Year 2: $XXM (X% of total)
   Year 3: $XXM (X% of total)
   Total 3-Year Synergies: $XXM

4. REALIZATION TIMELINE
   Immediate (0-6 months): [Which synergies + $XXM]
   Medium-term (6-18 months): [Which synergies + $XXM]
   Long-term (18-36 months): [Which synergies + $XXM]

5. RISK ASSESSMENT
   High Confidence (>80% probability): $XXM
   Medium Confidence (50-80%): $XXM
   Low Confidence (<50%): $XXM

6. NET IMPACT
   Gross Synergies: $XXM
   Integration Costs: ($XXM)
   Net Value Creation: $XXM
   NPV of Synergies: $XXM

Be specific with numbers. Show your math."""
    }

    base_prompt = prompts.get(module,
                              f"Analyze the {module} aspects of this M&A opportunity.")

    if documents:
        doc_section = "\n\nRELEVANT DOCUMENTS:\n" + "\n---\n".join(
            documents[:3])
        base_prompt += doc_section

    return base_prompt


# FastAPI app setup
app = FastAPI(title="CorpDev AI Platform", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Pydantic models
class AnalysisRequest(BaseModel):
    module: str
    acquirer: str = ""
    target: str = ""
    notes: str = ""
    targetIndustry: str = ""
    geography: str = ""
    revenueRange: str = ""
    wacc: str = ""
    terminalGrowth: str = ""
    taxRate: str = ""
    targetMarket: str = ""
    sampleCompany: str = ""
    upload_hashes: List[str] = Field(default_factory=list)
    api_key: str = ""


class ChatRequest(BaseModel):
    message: str
    api_key: str = ""
    context: str = ""


# API Routes
@app.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@app.get("/api/status")
async def get_status():
    return JSONResponse({
        "ok": True,
        "data": {
            "modules": app_state["modules"],
            "dependencies": {
                "openai": OPENAI_AVAILABLE,
                "pdf": PDF_AVAILABLE,
                "excel": XLSX_AVAILABLE,
                "pandas": PANDAS_AVAILABLE
            }
        }
    })


@app.post("/api/analyze")
async def run_analysis(request: AnalysisRequest):
    if not request.api_key:
        raise HTTPException(status_code=400, detail="OpenAI API key required")

    if request.module not in app_state["modules"]:
        raise HTTPException(status_code=400,
                            detail=f"Unknown module: {request.module}")

    # Gather document content (if any)
    documents = []
    for file_hash in request.upload_hashes:
        if file_hash in app_state["uploads"]:
            file_info = app_state["uploads"][file_hash]
            documents.append(
                f"File: {file_info['name']}\n{file_info['content']}")

    # Build form data dict
    form_data = {
        "acquirer": request.acquirer,
        "target": request.target,
        "notes": request.notes,
        "targetIndustry": request.targetIndustry,
        "geography": request.geography,
        "revenueRange": request.revenueRange,
        "wacc": request.wacc,
        "terminalGrowth": request.terminalGrowth,
        "taxRate": request.taxRate,
        "targetMarket": request.targetMarket,
        "sampleCompany": request.sampleCompany
    }

    # Build prompt
    prompt = build_analysis_prompt(request.module, form_data, documents)

    # Call OpenAI
    max_tokens = 3000 if request.module == "target_sourcing" else 2500
    result = call_openai_api(prompt, request.api_key, max_tokens=max_tokens)

    return JSONResponse({
        "ok": True,
        "data": {
            "analysis": result,
            "module": request.module,
            "timestamp": datetime.utcnow().isoformat()
        }
    })


@app.post("/api/chat")
async def chat(request: ChatRequest):
    if not request.api_key:
        raise HTTPException(status_code=400, detail="OpenAI API key required")

    prompt = f"""User question: {request.message}

Context from previous analysis:
{request.context[:2000]}

Provide a concise, helpful response focused on refining the analysis or answering specific questions."""

    result = call_openai_api(prompt, request.api_key, max_tokens=1000)

    return JSONResponse({
        "ok": True,
        "data": {"response": result}
    })


# Serve the HTML frontend
@app.get("/")
async def serve_index():
    return FileResponse("index.html")


if __name__ == "__main__":
    print("=" * 60)
    print("Starting Enhanced CorpDev AI Platform v2.0")
    print("=" * 60)
    print(f"Upload directory: {UPLOAD_DIR}")
    print(f"\nDependencies available:")
    print(f"  - OpenAI: {OPENAI_AVAILABLE}")
    print(f"  - PDF: {PDF_AVAILABLE}")
    print(f"  - Excel: {XLSX_AVAILABLE}")
    print(f"  - Pandas: {PANDAS_AVAILABLE}")
    print("\n" + "=" * 60)
    print("Server starting at http://127.0.0.1:8080")
    print("=" * 60 + "\n")

    uvicorn.run(
        "main:app",
        host="127.0.0.1",
        port=8080,
        reload=True,
        log_level="info"
    )